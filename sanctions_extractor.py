"""
═══════════════════════════════════════════════════════════════
  Global Sanctions Lists Extractor  —  VS Code / Terminal
  Full Column Edition
═══════════════════════════════════════════════════════════════

  Sources:
    1. OFAC   (US Treasury SDN)           → XML  treasury.gov
    2. UK HMT (OFSI Consolidated List)    → CSV  ofsistorage.blob.core.windows.net
    3. EU FSF (Financial Sanctions Files) → CSV  webgate.ec.europa.eu
    4. UNSC   (UN Security Council)       → XML  scsanctions.un.org

  Output:  sanctions_output.xlsx  (same folder as this script)

  Setup (run once):
    pip install requests openpyxl

  Run:
    python sanctions_extractor.py
═══════════════════════════════════════════════════════════════
"""

import csv
import re
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Sources ───────────────────────────────────────────────────────────────────

SOURCES = {
    'OFAC': {
        'url':      'https://www.treasury.gov/ofac/downloads/sdn.xml',
        'format':   'xml_ofac',
        'filename': 'ofac_sdn.xml',
    },
    'UK HMT': {
        'url':      'https://ofsistorage.blob.core.windows.net/publishlive/2022format/ConList.csv',
        'format':   'csv_uk',
        'filename': 'uk_conlist.csv',
    },
    'EU': {
        'url':      'https://webgate.ec.europa.eu/fsd/fsf/public/files/csvFullSanctionsList/content?token=dG9rZW4tMjAxNw',
        'format':   'csv_eu',
        'filename': 'eu_sanctions.csv',
    },
    'UNSC': {
        'url':      'https://scsanctions.un.org/resources/xml/en/consolidated.xml',
        'format':   'xml_unsc',
        'filename': 'unsc_consolidated.xml',
    },
}

ORDER       = ['OFAC', 'UK HMT', 'EU', 'UNSC']
SCRIPT_DIR  = Path(__file__).parent
CACHE_DIR   = SCRIPT_DIR / 'sanctions_cache'
OUTPUT_FILE = SCRIPT_DIR / 'sanctions_output.xlsx'
CACHE_DIR.mkdir(exist_ok=True)

# ── Master columns (ALL sources combined) ─────────────────────────────────────

MASTER_COLS = [
    'Source',
    'Reference / UID',
    'Sanctions Programme',
    'Last Name',
    'First Name',
    'Second / Middle Name',
    'Third Name',
    'Fourth Name',
    'Title',
    'Designation / Position',
    'Gender',
    'Good Quality AKA',
    'Low / Weak Quality AKA',
    'Date of Birth',
    'Place of Birth',
    'Country of Birth',
    'Nationality 1',
    'Nationality 2',
    'Passport Number',
    'Passport Country',
    'Passport Expiry',
    'National ID Number',
    'National ID Type',
    'National ID Country',
    'Address',
    'City',
    'Country',
    'Postal Code',
    'Listed On',
    'Last Updated',
    'Group Status',
    'Remarks / Other Info',
]

MASTER_COL_W = [
    10, 16, 28,
    28, 20, 20, 20, 20,
    14, 30, 10,
    40, 40,
    18, 24, 20,
    20, 20,
    20, 18, 16,
    22, 20, 18,
    36, 18, 18, 12,
    14, 14, 14,
    48,
]

COLORS = {
    'OFAC':   {'row': 'FFF2CC', 'header': '7F6000', 'sep': 'BF9000'},
    'UK HMT': {'row': 'D9EAF7', 'header': '1F4E79', 'sep': '2E75B6'},
    'EU':     {'row': 'E2F0D9', 'header': '375623', 'sep': '548235'},
    'UNSC':   {'row': 'FCE4D6', 'header': '843C0C', 'sep': 'C55A11'},
}

SOURCE_COLS = {
    'OFAC': [
        'Source', 'Reference / UID', 'Sanctions Programme',
        'Last Name', 'First Name', 'Title', 'Designation / Position',
        'Good Quality AKA', 'Low / Weak Quality AKA',
        'Date of Birth', 'Place of Birth', 'Nationality 1',
        'Passport Number', 'Passport Country', 'Passport Expiry',
        'National ID Number', 'National ID Type', 'National ID Country',
        'Address', 'City', 'Country', 'Postal Code',
        'Remarks / Other Info',
    ],
    'UK HMT': [
        'Source', 'Reference / UID', 'Sanctions Programme',
        'Last Name', 'First Name', 'Second / Middle Name',
        'Third Name', 'Fourth Name', 'Title', 'Designation / Position',
        'Good Quality AKA', 'Low / Weak Quality AKA',
        'Date of Birth', 'Place of Birth', 'Country of Birth', 'Nationality 1',
        'Passport Number', 'National ID Number',
        'Address', 'City', 'Country', 'Postal Code',
        'Listed On', 'Last Updated', 'Group Status',
        'Remarks / Other Info',
    ],
    'EU': [
        'Source', 'Reference / UID', 'Sanctions Programme',
        'Last Name', 'First Name', 'Second / Middle Name',
        'Title', 'Designation / Position', 'Gender',
        'Good Quality AKA',
        'Date of Birth', 'Place of Birth', 'Country of Birth',
        'Nationality 1', 'Nationality 2',
        'Passport Number', 'Passport Country', 'Passport Expiry',
        'National ID Number', 'National ID Type', 'National ID Country',
        'Address', 'City', 'Country', 'Postal Code',
        'Listed On',
    ],
    'UNSC': [
        'Source', 'Reference / UID', 'Sanctions Programme',
        'Last Name', 'First Name', 'Second / Middle Name',
        'Third Name', 'Fourth Name', 'Designation / Position',
        'Good Quality AKA', 'Low / Weak Quality AKA',
        'Date of Birth', 'Place of Birth', 'Country of Birth', 'Nationality 1',
        'Passport Number', 'Passport Country', 'Passport Expiry',
        'National ID Number', 'National ID Type', 'National ID Country',
        'Address', 'City', 'Country',
        'Listed On', 'Remarks / Other Info',
    ],
}

# ── Helpers ───────────────────────────────────────────────────────────────────

ILLEGAL_RE = re.compile(r'[\x00-\x08\x0b-\x0c\x0e-\x1f]')

def s(v):
    if not isinstance(v, str):
        return v or ''
    return ILLEGAL_RE.sub('', v).strip()

def empty_row(source):
    return {c: '' for c in MASTER_COLS} | {'Source': source}

# ── Downloader ────────────────────────────────────────────────────────────────

def download(key, cfg):
    dest = CACHE_DIR / cfg['filename']
    if dest.exists():
        print(f'  [{key}] Cached  ({dest.stat().st_size/1024/1024:.1f} MB)')
        return dest
    print(f'  [{key}] Downloading ...', end='', flush=True)
    r = requests.get(cfg['url'],
                     headers={'User-Agent': 'SanctionsExtractor/3.0'},
                     timeout=180, stream=True)
    r.raise_for_status()
    with open(dest, 'wb') as fh:
        for chunk in r.iter_content(65536):
            fh.write(chunk)
    print(f'  {dest.stat().st_size/1024/1024:.1f} MB  ✓')
    return dest

# ── Parsers ───────────────────────────────────────────────────────────────────

def parse_ofac_xml(path):
    rows = []
    try:
        tree = ET.parse(path)
        root = tree.getroot()
    except ET.ParseError as e:
        print(f'\n  ✗ OFAC XML error: {e}')
        return rows

    ns_m = re.match(r'\{(.*?)\}', root.tag)
    ns   = f'{{{ns_m.group(1)}}}' if ns_m else ''

    def tx(el, tag):
        n = el.find(f'{ns}{tag}')
        return n.text.strip() if n is not None and n.text else ''

    for entry in root.findall(f'.//{ns}sdnEntry'):
        if tx(entry, 'sdnType').lower() != 'individual':
            continue

        row = empty_row('OFAC')
        row['Reference / UID']      = s(tx(entry, 'uid'))
        row['Last Name']            = s(tx(entry, 'lastName'))
        row['First Name']           = s(tx(entry, 'firstName'))
        row['Title']                = s(tx(entry, 'title'))
        row['Remarks / Other Info'] = s(tx(entry, 'remarks'))

        progs = [s(p.text) for p in entry.findall(f'.//{ns}program') if p.text]
        row['Sanctions Programme'] = '; '.join(progs)

        good, weak = [], []
        for aka in entry.findall(f'.//{ns}aka'):
            cat  = tx(aka, 'category').lower()
            name = ' '.join(filter(None, [s(tx(aka, 'firstName')), s(tx(aka, 'lastName'))]))
            if name:
                (good if ('strong' in cat or 'good' in cat) else weak).append(name)
        row['Good Quality AKA']       = '; '.join(good)
        row['Low / Weak Quality AKA'] = '; '.join(weak)

        for dob_el in entry.findall(f'.//{ns}dateOfBirthItem'):
            dob = s(tx(dob_el, 'dateOfBirth'))
            if dob: row['Date of Birth'] = dob; break

        for pob_el in entry.findall(f'.//{ns}placeOfBirthItem'):
            pob = s(tx(pob_el, 'placeOfBirth'))
            if pob: row['Place of Birth'] = pob; break

        nats = [s(tx(n, 'country')) for n in entry.findall(f'.//{ns}nationalityItem') if tx(n, 'country')]
        if nats: row['Nationality 1'] = nats[0]
        if len(nats) > 1: row['Nationality 2'] = nats[1]

        passports, other_ids = [], []
        for id_el in entry.findall(f'.//{ns}id'):
            id_type = s(tx(id_el, 'idType'))
            id_num  = s(tx(id_el, 'idNumber'))
            if not id_num: continue
            if 'passport' in id_type.lower():
                passports.append((id_num, s(tx(id_el,'idCountry')), s(tx(id_el,'idExpirationDate'))))
            else:
                other_ids.append((id_type, id_num, s(tx(id_el,'idCountry'))))

        if passports:
            row['Passport Number']  = passports[0][0]
            row['Passport Country'] = passports[0][1]
            row['Passport Expiry']  = passports[0][2]
            if len(passports) > 1:
                extras = '; '.join(f'{p[0]} ({p[1]})' for p in passports[1:])
                row['Remarks / Other Info'] = (row['Remarks / Other Info'] + f' | Extra passports: {extras}').strip(' |')
        if other_ids:
            row['National ID Type']    = other_ids[0][0]
            row['National ID Number']  = other_ids[0][1]
            row['National ID Country'] = other_ids[0][2]

        for addr in entry.findall(f'.//{ns}address'):
            parts = [s(tx(addr, t)) for t in ['address1','address2','address3']]
            row['Address']     = '; '.join(p for p in parts if p)
            row['City']        = s(tx(addr, 'city'))
            row['Country']     = s(tx(addr, 'country'))
            row['Postal Code'] = s(tx(addr, 'postalCode'))
            break

        rows.append(row)
    return rows


def parse_uk(path):
    groups = {}
    with open(path, encoding='utf-8-sig', errors='replace', newline='') as fh:
        reader  = csv.reader(fh)
        headers = None
        for i, row in enumerate(reader):
            if i == 0: continue
            if i == 1: headers = [h.strip() for h in row]; continue
            if not headers or not any(row): continue
            d = dict(zip(headers, row + [''] * max(0, len(headers) - len(row))))
            if d.get('Group Type', '').strip() != 'Individual': continue
            gid = d.get('Group ID', '').strip()
            if gid not in groups:
                groups[gid] = {'data': d, 'aliases': [], 'address_done': False}
            grp = groups[gid]
            atype = d.get('Alias Type', '').strip()
            anames = [d.get(f'Alias Name {j}', '').strip() for j in range(1, 7)]
            aname  = ' '.join(a for a in anames if a)
            if aname:
                grp['aliases'].append((atype, aname))

    rows = []
    for gid, grp in groups.items():
        d   = grp['data']
        row = empty_row('UK HMT')
        row['Reference / UID']        = s(gid)
        row['Sanctions Programme']    = s(d.get('Regime', ''))
        row['Last Name']              = s(d.get('Name 6', ''))
        row['First Name']             = s(d.get('Name 1', ''))
        row['Second / Middle Name']   = s(d.get('Name 2', ''))
        row['Third Name']             = s(d.get('Name 3', ''))
        row['Fourth Name']            = s(d.get('Name 4', ''))
        row['Title']                  = s(d.get('Name 5', ''))
        row['Designation / Position'] = s(d.get('Designation', '') or d.get('Position', ''))
        row['Date of Birth']          = s(d.get('DOB', ''))
        row['Place of Birth']         = s(d.get('Town of Birth', ''))
        row['Country of Birth']       = s(d.get('Country of Birth', ''))
        row['Nationality 1']          = s(d.get('Nationality', ''))
        row['Passport Number']        = s(d.get('Passport Number', ''))
        row['National ID Number']     = s(d.get('National Identification Number', ''))
        row['Listed On']              = s(d.get('Listed On', '') or d.get('Individual Listed On', ''))
        row['Last Updated']           = s(d.get('Last Updated', ''))
        row['Group Status']           = s(d.get('Group Status', ''))
        row['Remarks / Other Info']   = s(d.get('Other Information', ''))

        addr_parts = [d.get(f'Address {j}', '').strip() for j in range(1, 7)]
        row['Address']     = '; '.join(a for a in addr_parts if a)
        row['Country']     = s(d.get('Country', ''))
        row['Postal Code'] = s(d.get('Post/Zip Code', ''))

        good = [name for atype, name in grp['aliases'] if not atype or 'primary' in atype.lower()]
        weak = [name for atype, name in grp['aliases'] if atype and 'primary' not in atype.lower()]
        row['Good Quality AKA']       = '; '.join(good)
        row['Low / Weak Quality AKA'] = '; '.join(weak)

        rows.append(row)
    return rows


def parse_eu(path):
    groups = {}
    with open(path, encoding='utf-8-sig', errors='replace', newline='') as fh:
        for line in fh:
            parts = line.rstrip('\r\n').split(';')
            if len(parts) < 18: continue
            if parts[2].strip() != 'P': continue
            eid = parts[1].strip()
            if eid not in groups:
                groups[eid] = {k: '' for k in [
                    'last','first','second','gender','title','function',
                    'dob','pob_city','pob_country','nat1','nat2',
                    'passport','passport_country','passport_expiry',
                    'nid','nid_type','nid_country',
                    'addr_street','addr_city','addr_country','addr_zip',
                    'programme','listed_on','ref',
                ]}
                groups[eid]['aliases'] = []
            g = groups[eid]

            def col(i): return parts[i].strip() if len(parts) > i else ''

            if not g['last']    and col(14): g['last']    = col(14)
            if not g['first']   and col(15): g['first']   = col(15)
            if not g['second']  and col(16): g['second']  = col(16)
            if not g['gender']  and col(17): g['gender']  = col(17)
            if not g['title']   and col(18): g['title']   = col(18)
            if not g['function']and col(19): g['function']= col(19)
            if not g['dob']     and col(36): g['dob']     = col(36)
            if not g['dob']     and col(20): g['dob']     = col(20)
            if not g['pob_city']    and col(21): g['pob_city']    = col(21)
            if not g['pob_country'] and col(23): g['pob_country'] = col(23)
            if not g['nat1']    and col(33): g['nat1']    = col(33)
            if not g['nat2']    and col(35): g['nat2']    = col(35)
            if not g['passport']         and col(43): g['passport']         = col(43)
            if not g['passport_country'] and col(44): g['passport_country'] = col(44)
            if not g['passport_expiry']  and col(46): g['passport_expiry']  = col(46)
            if not g['nid']         and col(38): g['nid']         = col(38)
            if not g['nid_type']    and col(37): g['nid_type']    = col(37)
            if not g['nid_country'] and col(39): g['nid_country'] = col(39)
            if not g['addr_street']  and col(29): g['addr_street']  = col(29)
            if not g['addr_city']    and col(26): g['addr_city']    = col(26)
            if not g['addr_country'] and col(27): g['addr_country'] = col(27)
            if not g['addr_zip']     and col(31): g['addr_zip']     = col(31)
            alias = ' '.join(filter(None, [col(24), col(25)]))
            if alias and alias not in g['aliases']: g['aliases'].append(alias)
            if not g['programme'] and col(9):  g['programme'] = col(9)
            if not g['listed_on'] and col(5):  g['listed_on'] = col(5)
            if not g['ref']       and col(3):  g['ref']       = col(3)

    rows = []
    for eid, g in groups.items():
        if not g['last'] and not g['first']: continue
        row = empty_row('EU')
        row['Reference / UID']        = s(g['ref'] or eid)
        row['Sanctions Programme']    = s(g['programme'])
        row['Last Name']              = s(g['last'])
        row['First Name']             = s(g['first'])
        row['Second / Middle Name']   = s(g['second'])
        row['Title']                  = s(g['title'])
        row['Designation / Position'] = s(g['function'])
        row['Gender']                 = s(g['gender'])
        row['Good Quality AKA']       = '; '.join(g['aliases'])
        row['Date of Birth']          = s(g['dob'])
        row['Place of Birth']         = s(g['pob_city'])
        row['Country of Birth']       = s(g['pob_country'])
        row['Nationality 1']          = s(g['nat1'])
        row['Nationality 2']          = s(g['nat2'])
        row['Passport Number']        = s(g['passport'])
        row['Passport Country']       = s(g['passport_country'])
        row['Passport Expiry']        = s(g['passport_expiry'])
        row['National ID Number']     = s(g['nid'])
        row['National ID Type']       = s(g['nid_type'])
        row['National ID Country']    = s(g['nid_country'])
        row['Address']                = s(g['addr_street'])
        row['City']                   = s(g['addr_city'])
        row['Country']                = s(g['addr_country'])
        row['Postal Code']            = s(g['addr_zip'])
        row['Listed On']              = s(g['listed_on'])
        rows.append(row)
    return rows


def parse_unsc(path):
    rows = []
    try:
        tree = ET.parse(path)
        root = tree.getroot()
    except ET.ParseError as e:
        print(f'\n  ✗ UNSC XML error: {e}')
        return rows

    ns_m = re.match(r'\{(.*?)\}', root.tag)
    ns   = f'{{{ns_m.group(1)}}}' if ns_m else ''

    def tx(el, tag):
        n = el.find(f'{ns}{tag}')
        return n.text.strip() if n is not None and n.text else ''

    for indiv in root.findall(f'.//{ns}INDIVIDUAL'):
        row = empty_row('UNSC')
        row['Reference / UID']        = s(tx(indiv, 'REFERENCE_NUMBER'))
        row['Sanctions Programme']    = s(tx(indiv, 'UN_LIST_TYPE'))
        row['First Name']             = s(tx(indiv, 'FIRST_NAME'))
        row['Second / Middle Name']   = s(tx(indiv, 'SECOND_NAME'))
        row['Third Name']             = s(tx(indiv, 'THIRD_NAME'))
        row['Fourth Name']            = s(tx(indiv, 'FOURTH_NAME'))
        row['Last Name']              = s(tx(indiv, 'FOURTH_NAME') or tx(indiv, 'THIRD_NAME'))
        row['Listed On']              = s(tx(indiv, 'LISTED_ON'))
        row['Remarks / Other Info']   = s(tx(indiv, 'COMMENTS1'))

        # Designation
        desig_el = indiv.find(f'.//{ns}DESIGNATION/{ns}VALUE')
        if desig_el is not None and desig_el.text:
            row['Designation / Position'] = s(desig_el.text)

        # Nationality
        nats = [s(tx(n, 'VALUE')) for n in indiv.findall(f'.//{ns}NATIONALITY') if tx(n, 'VALUE')]
        if nats: row['Nationality 1'] = nats[0]
        if len(nats) > 1: row['Nationality 2'] = nats[1]

        # Aliases
        good, weak = [], []
        for alias in indiv.findall(f'{ns}INDIVIDUAL_ALIAS'):
            quality = tx(alias, 'QUALITY').lower()
            name    = s(tx(alias, 'ALIAS_NAME'))
            if not name: continue
            (good if 'good' in quality else weak).append(name)
        row['Good Quality AKA']       = '; '.join(good)
        row['Low / Weak Quality AKA'] = '; '.join(weak)

        # DOB
        for dob_el in indiv.findall(f'{ns}INDIVIDUAL_DATE_OF_BIRTH'):
            yr = tx(dob_el, 'YEAR')
            mo = tx(dob_el, 'MONTH') or tx(dob_el, 'FROM_MONTH')
            dy = tx(dob_el, 'DAY')   or tx(dob_el, 'FROM_DAY')
            if yr:
                row['Date of Birth'] = f'{dy}/{mo}/{yr}'.strip('/') if (mo or dy) else yr
                break

        # POB
        for pob_el in indiv.findall(f'{ns}INDIVIDUAL_PLACE_OF_BIRTH'):
            row['Place of Birth']   = s(tx(pob_el, 'CITY') or tx(pob_el, 'STATE_PROVINCE'))
            row['Country of Birth'] = s(tx(pob_el, 'COUNTRY'))
            break

        # Documents
        passports, other_ids = [], []
        for doc in indiv.findall(f'{ns}INDIVIDUAL_DOCUMENT'):
            doc_type = tx(doc, 'TYPE_OF_DOCUMENT')
            doc_num  = s(tx(doc, 'NUMBER'))
            if not doc_num: continue
            entry_doc = (doc_num, s(tx(doc,'ISSUING_COUNTRY')), s(tx(doc,'EXPIRY_DATE')), s(doc_type))
            if 'passport' in doc_type.lower():
                passports.append(entry_doc)
            else:
                other_ids.append(entry_doc)

        if passports:
            row['Passport Number']  = passports[0][0]
            row['Passport Country'] = passports[0][1]
            row['Passport Expiry']  = passports[0][2]
        if other_ids:
            row['National ID Number']  = other_ids[0][0]
            row['National ID Country'] = other_ids[0][1]
            row['National ID Type']    = other_ids[0][3]

        # Address
        for addr in indiv.findall(f'{ns}INDIVIDUAL_ADDRESS'):
            parts = [s(tx(addr, t)) for t in ['STREET', 'STATE_PROVINCE', 'NOTE']]
            row['Address'] = '; '.join(p for p in parts if p)
            row['City']    = s(tx(addr, 'CITY'))
            row['Country'] = s(tx(addr, 'COUNTRY'))
            break

        rows.append(row)
    return rows


PARSERS = {
    'xml_ofac': parse_ofac_xml,
    'csv_uk':   parse_uk,
    'csv_eu':   parse_eu,
    'xml_unsc': parse_unsc,
}

# ── Excel builder ─────────────────────────────────────────────────────────────

def build_excel(all_rows, counts):
    _thin   = Side(style='thin', color='D0D0D0')
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

    def hdr_cell(cell, fg, text):
        cell.value     = text
        cell.font      = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor=fg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = _border

    def data_cell(cell, val, fg):
        cell.value     = s(str(val)) if val else ''
        cell.font      = Font(name='Arial', size=9)
        cell.fill      = PatternFill('solid', fgColor=fg)
        cell.alignment = Alignment(vertical='top', wrap_text=False)
        cell.border    = _border

    wb = Workbook()

    # ── Sheet 1: All Records ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'All Records'
    for ci, (h, w) in enumerate(zip(MASTER_COLS, MASTER_COL_W), 1):
        hdr_cell(ws.cell(row=1, column=ci), '1F3864', h)
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 30

    row_num, cur_source = 2, None
    for entry in all_rows:
        src = entry['Source']
        if src != cur_source:
            for ci in range(1, len(MASTER_COLS) + 1):
                c = ws.cell(row=row_num, column=ci,
                            value=f'  ▶  {src}' if ci == 1 else '')
                c.font      = Font(name='Arial', bold=True, size=10, color='FFFFFF')
                c.fill      = PatternFill('solid', fgColor=COLORS[src]['sep'])
                c.alignment = Alignment(vertical='center')
                c.border    = _border
            ws.row_dimensions[row_num].height = 16
            row_num += 1; cur_source = src
        for ci, col in enumerate(MASTER_COLS, 1):
            data_cell(ws.cell(row=row_num, column=ci), entry.get(col, ''), COLORS[src]['row'])
        ws.row_dimensions[row_num].height = 15
        row_num += 1

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(MASTER_COLS))}1'

    # ── Per-source sheets ─────────────────────────────────────────────────────
    for src in ORDER:
        src_rows = [r for r in all_rows if r['Source'] == src]
        cols     = SOURCE_COLS[src]
        ws_s     = wb.create_sheet(src.replace(' ', '_'))
        color    = COLORS[src]
        for ci, col in enumerate(cols, 1):
            idx = MASTER_COLS.index(col) if col in MASTER_COLS else -1
            w   = MASTER_COL_W[idx] if idx >= 0 else 20
            hdr_cell(ws_s.cell(row=1, column=ci), color['header'], col)
            ws_s.column_dimensions[get_column_letter(ci)].width = w
        ws_s.row_dimensions[1].height = 30
        for ri, entry in enumerate(src_rows, 2):
            for ci, col in enumerate(cols, 1):
                data_cell(ws_s.cell(row=ri, column=ci), entry.get(col, ''), color['row'])
            ws_s.row_dimensions[ri].height = 15
        ws_s.freeze_panes = 'A2'
        ws_s.auto_filter.ref = f'A1:{get_column_letter(len(cols))}1'

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Summary', 1)
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 52

    ws2.merge_cells('A1:C1')
    ws2['A1'].value     = 'Global Sanctions Lists — Full Column Extraction'
    ws2['A1'].font      = Font(name='Arial', bold=True, size=14, color='1F3864')
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 32

    ws2.merge_cells('A2:C2')
    ws2['A2'].value     = f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M')}"
    ws2['A2'].font      = Font(name='Arial', italic=True, size=10, color='808080')
    ws2['A2'].alignment = Alignment(horizontal='center')

    for ci, h in enumerate(['Source', 'Records', 'Columns Captured'], 1):
        hdr_cell(ws2.cell(row=4, column=ci), '1F3864', h)
    ws2.row_dimensions[4].height = 24

    for ri, src in enumerate(ORDER, 5):
        col_list = ', '.join(c for c in SOURCE_COLS[src] if c != 'Source')
        c1 = ws2.cell(row=ri, column=1, value=src)
        c2 = ws2.cell(row=ri, column=2, value=counts.get(src, 0))
        c3 = ws2.cell(row=ri, column=3, value=col_list)
        fill = PatternFill('solid', fgColor=COLORS[src]['row'])
        for c, ha in [(c1,'left'),(c2,'center'),(c3,'left')]:
            c.font      = Font(name='Arial', size=10, bold=(c == c1))
            c.fill      = fill
            c.alignment = Alignment(horizontal=ha, vertical='top', wrap_text=True)
            c.border    = _border
        ws2.row_dimensions[ri].height = 40

    total_row = 5 + len(ORDER)
    for c, v, ha in [
        (ws2.cell(row=total_row, column=1), 'TOTAL', 'left'),
        (ws2.cell(row=total_row, column=2), sum(counts.values()), 'center'),
        (ws2.cell(row=total_row, column=3), f'{len(MASTER_COLS)} master columns across all sources', 'left'),
    ]:
        c.value     = v
        c.font      = Font(name='Arial', bold=True, size=11)
        c.alignment = Alignment(horizontal=ha, vertical='center')
    ws2.row_dimensions[total_row].height = 20

    wb.save(OUTPUT_FILE)

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print('=' * 62)
    print('  Global Sanctions Lists Extractor — Full Column Edition')
    print(f'  {datetime.now().strftime("%d %B %Y  %H:%M")}')
    print('=' * 62)

    print('\n[1/3] Downloading lists...\n')
    local_files = {}
    for key, cfg in SOURCES.items():
        try:
            local_files[key] = download(key, cfg)
        except Exception as e:
            print(f'\n  [{key}] FAILED: {e}')
            local_files[key] = None

    print('\n[2/3] Parsing...\n')
    all_rows, counts = [], {}
    for key in ORDER:
        path = local_files.get(key)
        if not path or not path.exists():
            print(f'  [{key}] Skipped (no file)')
            counts[key] = 0
            continue
        fmt     = SOURCES[key]['format']
        size_mb = path.stat().st_size / 1024 / 1024
        print(f'  [{key}] {path.name} ({size_mb:.1f} MB) ...', end='', flush=True)
        try:
            rows = PARSERS[fmt](path)
        except Exception as e:
            print(f'\n  ✗ Error: {e}')
            rows = []
        counts[key] = len(rows)
        all_rows.extend(rows)
        print(f'  {len(rows):,} records  ✓')

    print(f'\n  Total: {len(all_rows):,} records')

    print('\n[3/3] Building Excel...')
    build_excel(all_rows, counts)

    print(f'\n✓ Saved → {OUTPUT_FILE}')
    print(f'\n  {"Source":<12}  {"Records":>9}  Columns')
    print(f'  {"-"*12}  {"-"*9}  {"-"*30}')
    for src in ORDER:
        print(f'  {src:<12}  {counts.get(src,0):>9,}  {len(SOURCE_COLS[src])} columns')
    print(f'  {"TOTAL":<12}  {sum(counts.values()):>9,}  {len(MASTER_COLS)} in combined sheet')
    print(f'\n  Sheets : All Records | Summary | OFAC | UK_HMT | EU | UNSC')
    print(f'  Cache  : {CACHE_DIR}')
    print(f'           (delete this folder to force a fresh download)')

if __name__ == '__main__':
    main()
